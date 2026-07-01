from pathlib import Path

import pytest
from openpyxl import Workbook

from job_radar.config import ConfigError
from job_radar.job_history import EXPECTED_HEADERS, load_job_history_workbook


def write_history_workbook(
    workbook_path: Path,
    pipeline_rows: list[list[object]],
    reviewed_rows: list[list[object]],
    headers: list[str] | None = None,
) -> None:
    workbook = Workbook()
    pipeline_sheet = workbook.active
    pipeline_sheet.title = "Pipeline Import"
    reviewed_sheet = workbook.create_sheet("Reviewed Import")

    sheet_headers = headers or EXPECTED_HEADERS

    pipeline_sheet.append(sheet_headers)
    reviewed_sheet.append(sheet_headers)

    for row in pipeline_rows:
        pipeline_sheet.append(row)

    for row in reviewed_rows:
        reviewed_sheet.append(row)

    workbook.save(workbook_path)


def make_history_row(
    history_type: str = "pipeline",
    company: str = "Example AI",
    role: str = "Senior Infrastructure Engineer",
    source: str = "LinkedIn",
    ats_platform: str = "Greenhouse",
    work_arrangement: str = "Remote",
    location: str = "Remote",
    comp_range: str = "$160k-$200k",
    event_date: str = "2026-06-01",
    status: str = "Rejected - No Interview",
    outcome_category: str = "Rejected No Interview",
    recruiter_contact: str = "Unknown",
    technical_match: str = "Very Strong",
    hiring_probability: str = "Low",
    skills_signals: str = "Linux, HPC, Infrastructure",
    primary_blocker: str = "Generic Remote Competition",
    secondary_blocker: str = "None",
    revisit: str = "No",
    include_in_job_radar: str = "Yes",
    import_key: str = "pipeline:example-ai:senior-infrastructure-engineer",
    notes: str = "Form rejection.",
) -> list[object]:
    return [
        history_type,
        company,
        role,
        source,
        ats_platform,
        work_arrangement,
        location,
        comp_range,
        event_date,
        status,
        outcome_category,
        recruiter_contact,
        technical_match,
        hiring_probability,
        skills_signals,
        primary_blocker,
        secondary_blocker,
        revisit,
        include_in_job_radar,
        import_key,
        notes,
    ]


def test_load_job_history_workbook_reads_pipeline_and_reviewed_rows(
    tmp_path: Path,
) -> None:
    workbook_path = tmp_path / "job-history.xlsx"

    write_history_workbook(
        workbook_path=workbook_path,
        pipeline_rows=[make_history_row()],
        reviewed_rows=[
            make_history_row(
                history_type="reviewed",
                company="SkipCo",
                role="Frontend Engineer",
                status="Skipped",
                outcome_category="Skipped",
                technical_match="Weak",
                hiring_probability="Very Low",
                primary_blocker="Role Family Mismatch",
                import_key="reviewed:skipco:frontend-engineer",
                notes="Not an infrastructure role.",
            )
        ],
    )

    result = load_job_history_workbook(workbook_path)

    assert result.rows_read == 2
    assert result.rows_imported == 2
    assert result.rows_skipped == 0
    assert len(result.records) == 2

    pipeline_record = result.records[0]
    reviewed_record = result.records[1]

    assert pipeline_record.history_type == "pipeline"
    assert pipeline_record.company == "Example AI"
    assert pipeline_record.role == "Senior Infrastructure Engineer"
    assert pipeline_record.event_date == "2026-06-01"
    assert pipeline_record.include_in_job_radar is True
    assert pipeline_record.import_key == "pipeline:example-ai:senior-infrastructure-engineer"

    assert reviewed_record.history_type == "reviewed"
    assert reviewed_record.company == "SkipCo"
    assert reviewed_record.primary_blocker == "Role Family Mismatch"
    assert reviewed_record.import_key == "reviewed:skipco:frontend-engineer"


def test_load_job_history_workbook_skips_rows_excluded_from_job_radar(
    tmp_path: Path,
) -> None:
    workbook_path = tmp_path / "job-history.xlsx"

    write_history_workbook(
        workbook_path=workbook_path,
        pipeline_rows=[
            make_history_row(
                include_in_job_radar="No",
                import_key="pipeline:excluded:role",
            )
        ],
        reviewed_rows=[],
    )

    result = load_job_history_workbook(workbook_path)

    assert result.rows_read == 1
    assert result.rows_imported == 0
    assert result.rows_skipped == 1
    assert result.records == []


def test_load_job_history_workbook_ignores_blank_rows(tmp_path: Path) -> None:
    workbook_path = tmp_path / "job-history.xlsx"

    write_history_workbook(
        workbook_path=workbook_path,
        pipeline_rows=[make_history_row(), [None] * len(EXPECTED_HEADERS)],
        reviewed_rows=[],
    )

    result = load_job_history_workbook(workbook_path)

    assert result.rows_read == 1
    assert result.rows_imported == 1
    assert result.rows_skipped == 0


def test_load_job_history_workbook_requires_expected_headers(tmp_path: Path) -> None:
    workbook_path = tmp_path / "job-history.xlsx"
    bad_headers = EXPECTED_HEADERS.copy()
    bad_headers[0] = "Wrong Header"

    write_history_workbook(
        workbook_path=workbook_path,
        pipeline_rows=[],
        reviewed_rows=[],
        headers=bad_headers,
    )

    with pytest.raises(ConfigError, match="headers do not match"):
        load_job_history_workbook(workbook_path)


def test_load_job_history_workbook_requires_import_key(tmp_path: Path) -> None:
    workbook_path = tmp_path / "job-history.xlsx"
    row = make_history_row()
    row[19] = None

    write_history_workbook(
        workbook_path=workbook_path,
        pipeline_rows=[row],
        reviewed_rows=[],
    )

    with pytest.raises(ConfigError, match="missing Import Key"):
        load_job_history_workbook(workbook_path)


def test_load_job_history_workbook_requires_both_import_sheets(tmp_path: Path) -> None:
    workbook_path = tmp_path / "job-history.xlsx"

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Pipeline Import"
    worksheet.append(EXPECTED_HEADERS)
    workbook.save(workbook_path)

    with pytest.raises(ConfigError, match="missing sheet: Reviewed Import"):
        load_job_history_workbook(workbook_path)