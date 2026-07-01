from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from job_radar.config import ConfigError
from job_radar.normalize import clean_text


PIPELINE_SHEET_NAME = "Pipeline Import"
REVIEWED_SHEET_NAME = "Reviewed Import"

EXPECTED_HEADERS = [
    "History Type",
    "Company",
    "Role",
    "Source",
    "ATS Platform",
    "Work Arrangement",
    "Location",
    "Comp Range",
    "Event Date",
    "Status",
    "Outcome Category",
    "Recruiter/Contact",
    "Technical Match",
    "Hiring Probability",
    "Skills/Signals",
    "Primary Blocker",
    "Secondary Blocker",
    "Revisit",
    "Include In Job Radar",
    "Import Key",
    "Notes",
]


@dataclass(frozen=True)
class JobHistoryRecord:
    history_type: str
    company: str
    role: str
    source: str | None
    ats_platform: str | None
    work_arrangement: str | None
    location: str | None
    comp_range: str | None
    event_date: str | None
    status: str | None
    outcome_category: str | None
    recruiter_contact: str | None
    technical_match: str | None
    hiring_probability: str | None
    skills_signals: str | None
    primary_blocker: str | None
    secondary_blocker: str | None
    revisit: str | None
    include_in_job_radar: bool
    import_key: str
    notes: str | None


@dataclass(frozen=True)
class JobHistoryImportResult:
    rows_read: int
    rows_imported: int
    rows_skipped: int
    records: list[JobHistoryRecord]


def load_job_history_workbook(path: str | Path) -> JobHistoryImportResult:
    workbook_path = Path(path)

    if not workbook_path.exists():
        raise ConfigError(f"Job history workbook does not exist: {workbook_path}")

    workbook = load_workbook(workbook_path, data_only=True)

    records: list[JobHistoryRecord] = []
    rows_read = 0
    rows_skipped = 0

    for sheet_name in [PIPELINE_SHEET_NAME, REVIEWED_SHEET_NAME]:
        if sheet_name not in workbook.sheetnames:
            raise ConfigError(f"Job history workbook is missing sheet: {sheet_name}")

        sheet_records, sheet_rows_read, sheet_rows_skipped = _load_history_sheet(
            workbook[sheet_name],
            sheet_name,
        )

        records.extend(sheet_records)
        rows_read += sheet_rows_read
        rows_skipped += sheet_rows_skipped

    return JobHistoryImportResult(
        rows_read=rows_read,
        rows_imported=len(records),
        rows_skipped=rows_skipped,
        records=records,
    )


def _load_history_sheet(
    worksheet: Any,
    sheet_name: str,
) -> tuple[list[JobHistoryRecord], int, int]:
    headers = [_cell_to_text(cell.value) for cell in worksheet[1]]

    if headers[: len(EXPECTED_HEADERS)] != EXPECTED_HEADERS:
        raise ConfigError(
            f"{sheet_name} headers do not match expected Job Radar history schema"
        )

    records: list[JobHistoryRecord] = []
    rows_read = 0
    rows_skipped = 0

    for row_number, row in enumerate(
        worksheet.iter_rows(min_row=2, max_col=len(EXPECTED_HEADERS)),
        start=2,
    ):
        values = {
            header: _cell_to_text(cell.value)
            for header, cell in zip(EXPECTED_HEADERS, row)
        }

        if _is_blank_row(values):
            continue

        rows_read += 1

        if _is_excluded(values["Include In Job Radar"]):
            rows_skipped += 1
            continue

        records.append(_build_record(values, sheet_name, row_number))

    return records, rows_read, rows_skipped


def _build_record(
    values: dict[str, str | None],
    sheet_name: str,
    row_number: int,
) -> JobHistoryRecord:
    history_type = _required_value(values["History Type"], sheet_name, row_number, "History Type")
    company = _required_value(values["Company"], sheet_name, row_number, "Company")
    role = _required_value(values["Role"], sheet_name, row_number, "Role")
    import_key = _required_value(values["Import Key"], sheet_name, row_number, "Import Key")

    return JobHistoryRecord(
        history_type=history_type,
        company=company,
        role=role,
        source=values["Source"],
        ats_platform=values["ATS Platform"],
        work_arrangement=values["Work Arrangement"],
        location=values["Location"],
        comp_range=values["Comp Range"],
        event_date=_normalize_event_date(values["Event Date"]),
        status=values["Status"],
        outcome_category=values["Outcome Category"],
        recruiter_contact=values["Recruiter/Contact"],
        technical_match=values["Technical Match"],
        hiring_probability=values["Hiring Probability"],
        skills_signals=values["Skills/Signals"],
        primary_blocker=values["Primary Blocker"],
        secondary_blocker=values["Secondary Blocker"],
        revisit=values["Revisit"],
        include_in_job_radar=True,
        import_key=import_key,
        notes=values["Notes"],
    )


def _required_value(
    value: str | None,
    sheet_name: str,
    row_number: int,
    column_name: str,
) -> str:
    if value is None:
        raise ConfigError(f"{sheet_name} row {row_number} is missing {column_name}")

    return value


def _normalize_event_date(value: str | None) -> str | None:
    if value is None:
        return None

    cleaned_value = clean_text(value)

    if not cleaned_value:
        return None

    return cleaned_value


def _cell_to_text(value: object) -> str | None:
    if value is None:
        return None

    if isinstance(value, datetime):
        return value.date().isoformat()

    if isinstance(value, date):
        return value.isoformat()

    cleaned_value = clean_text(str(value))

    if not cleaned_value:
        return None

    return cleaned_value


def _is_blank_row(values: dict[str, str | None]) -> bool:
    return all(value is None for value in values.values())


def _is_excluded(value: str | None) -> bool:
    return value is not None and value.lower() == "no"